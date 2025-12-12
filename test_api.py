#!/usr/bin/env python3
"""测试API端点是否可用"""
import requests
import json
import sys

BASE_URL = "http://localhost:5000"

def test_upload():
    """测试上传API"""
    print("测试 /api/upload...")
    try:
        # 创建一个简单的测试文件
        import tempfile
        import openpyxl
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '反馈'
        ws['A2'] = '测试反馈内容'
        ws['B1'] = '其他列'
        ws['B2'] = '其他内容'
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        with open(tmp_path, 'rb') as f:
            files = {'file': ('test.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(f"{BASE_URL}/api/upload", files=files, timeout=10)
        
        import os
        os.unlink(tmp_path)
        
        if response.status_code == 200:
            data = response.json()
            print(f"  ✓ 上传成功: fileId={data.get('fileId')}")
            return data.get('fileId')
        else:
            print(f"  ✗ 上传失败: {response.status_code} - {response.text}")
            return None
    except requests.exceptions.ConnectionError:
        print(f"  ✗ 无法连接到服务器 {BASE_URL}，请确保后端已启动")
        return None
    except Exception as e:
        print(f"  ✗ 上传测试失败: {e}")
        return None

def test_analyze(file_id):
    """测试分析API"""
    if not file_id:
        print("跳过分析测试（没有fileId）")
        return
    
    print(f"\n测试 /api/analyze (fileId={file_id})...")
    try:
        response = requests.post(
            f"{BASE_URL}/api/analyze",
            json={'fileId': file_id},
            headers={'Content-Type': 'application/json'},
            stream=True,
            timeout=5
        )
        
        if response.status_code == 200:
            print(f"  ✓ 分析请求已接受，状态码: {response.status_code}")
            print(f"  ✓ Content-Type: {response.headers.get('Content-Type')}")
            
            # 尝试读取前几个SSE消息
            print("  正在接收SSE流...")
            count = 0
            for line in response.iter_lines():
                if line:
                    line_str = line.decode('utf-8')
                    if line_str.startswith('data: '):
                        try:
                            data = json.loads(line_str[6:])
                            count += 1
                            if data.get('type') == 'progress':
                                print(f"  → 进度: {data.get('message')} ({data.get('current')}/{data.get('total')})")
                            elif data.get('type') == 'complete':
                                print(f"  ✓ 分析完成！收到 {len(data.get('data', {}).get('sheets', []))} 个sheet")
                                return True
                            elif data.get('type') == 'error':
                                print(f"  ✗ 分析错误: {data.get('message')}")
                                return False
                        except json.JSONDecodeError as e:
                            print(f"  ⚠ JSON解析失败: {e}")
                    
                    if count >= 10:  # 只读取前10条消息
                        print(f"  → 已收到 {count} 条消息，流仍在继续...")
                        break
        else:
            print(f"  ✗ 分析请求失败: {response.status_code} - {response.text[:200]}")
            return False
    except requests.exceptions.Timeout:
        print(f"  ⚠ 请求超时（5秒），但API可能仍在处理中")
        return None
    except requests.exceptions.ConnectionError:
        print(f"  ✗ 无法连接到服务器")
        return False
    except Exception as e:
        print(f"  ✗ 分析测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_stop(file_id):
    """测试停止API"""
    if not file_id:
        print("跳过停止测试（没有fileId）")
        return
    
    print(f"\n测试 /api/analyze/stop (fileId={file_id})...")
    try:
        response = requests.post(
            f"{BASE_URL}/api/analyze/stop",
            json={'fileId': file_id},
            headers={'Content-Type': 'application/json'},
            timeout=5
        )
        
        if response.status_code == 200:
            data = response.json()
            print(f"  ✓ 停止请求成功: {data.get('message')}")
            return True
        else:
            print(f"  ✗ 停止请求失败: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        print(f"  ✗ 停止测试失败: {e}")
        return False

def main():
    print("=" * 60)
    print("API可用性测试")
    print("=" * 60)
    
    # 测试上传
    file_id = test_upload()
    
    if file_id:
        # 测试分析（只测试几秒钟）
        test_analyze(file_id)
        
        # 测试停止
        test_stop(file_id)
    
    print("\n" + "=" * 60)
    print("测试完成")
    print("=" * 60)

if __name__ == '__main__':
    main()
