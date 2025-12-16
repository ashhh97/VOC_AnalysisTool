#!/usr/bin/env python3
from flask import Flask, request, jsonify, Response, stream_with_context
from flask_cors import CORS
import os
import uuid
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import threading
import time
import queue
from voc_analyzer import VOCAnalyzer

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

analyzer = VOCAnalyzer()

# 用于跟踪分析任务的状态
# 用于跟踪分析任务的状态
analysis_tasks = {}  # {file_id: {'stop_flag': threading.Event(), 'thread': thread}}

@app.route('/api/log_feedback', methods=['POST'])
def log_feedback():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Log to file
        log_file = os.path.join(os.path.dirname(__file__), 'training_data.jsonl')
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(json.dumps(data, ensure_ascii=False) + '\n')
            
        print(f"[Feedback] Logged implicit feedback: {data.get('event_type')}")
        return jsonify({'status': 'success', 'message': 'Feedback logged'}), 200
    except Exception as e:
        print(f"[Feedback] Error logging feedback: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/recalculate_stats', methods=['POST'])
def recalculate_stats():
    """重新计算统计数据（用户手动归类后调用）"""
    try:
        data = request.json
        if not data or 'celldata' not in data:
            return jsonify({'error': 'No celldata provided'}), 400
            
        celldata = data['celldata']
        print(f"[Recalculate] Received {len(celldata)} cells")
        
        # 解析表格数据
        # 假设格式: 行0是表头, 列0=问题概括, 列1=用户情绪, 列2=VOC原声
        rows_data = {}
        for cell in celldata:
            r, c = cell['r'], cell['c']
            if r == 0:  # 跳过表头
                continue
            if r not in rows_data:
                rows_data[r] = {}
            rows_data[r][c] = cell['v']['v']
        
        # 统计: 按(问题概括, 用户情绪)分组
        groups = {}
        total_rows = len(rows_data)
        
        for row_idx, row in rows_data.items():
            summary = row.get(0, '未分类')
            sentiment = row.get(1, '中性😐')
            snippet = row.get(2, '')
            
            key = (summary, sentiment)
            if key not in groups:
                groups[key] = {
                    'summary': summary,
                    'sentiment': sentiment,
                    'user_count': 0,
                    'snippets': []
                }
            groups[key]['user_count'] += 1
            groups[key]['snippets'].append(snippet)
        
        # 计算百分比并排序
        result_list = []
        for key, group in groups.items():
            user_pct = (group['user_count'] / total_rows * 100) if total_rows > 0 else 0
            result_list.append({
                'summary': group['summary'],
                'sentiment': group['sentiment'],
                'user_count': group['user_count'],
                'user_pct': f"{user_pct:.2f}%",
                'snippets': group['snippets']
            })
        
        # 按用户数量降序排序
        result_list.sort(key=lambda x: x['user_count'], reverse=True)
        
        # 构建新的celldata（带统计列）
        new_celldata = []
        
        # 表头
        headers = ['问题概括', '用户情绪', '用户数量', '用户占比', 'VOC原声片段']
        for i, header in enumerate(headers):
            new_celldata.append({
                'r': 0,
                'c': i,
                'v': {
                    'v': header,
                    'm': header,
                    'ct': {'fa': 'General', 't': 'g'},
                    'bg': '#EDEBE9',
                    'bl': 1
                }
            })
        
        current_row = 1
        merge_config = {}
        
        # 填充数据
        for group in result_list:
            start_row = current_row
            rows_count = len(group['snippets'])
            
            # 每个snippet一行
            for snippet in group['snippets']:
                new_celldata.append({
                    'r': current_row,
                    'c': 4,  # VOC原声片段
                    'v': {
                        'v': snippet,
                        'm': str(snippet),
                        'ct': {'fa': 'General', 't': 'g'}
                    }
                })
                current_row += 1
            
            # 统计列（合并单元格）
            # 问题概括
            new_celldata.append({
                'r': start_row,
                'c': 0,
                'v': {
                    'v': group['summary'],
                    'm': group['summary'],
                    'ct': {'fa': 'General', 't': 'g'},
                    'vt': 1, 'ht': 1,
                    'bg': '#E6F2FF'
                }
            })
            
            # 用户情绪（带颜色）
            font_color = '#000000'
            if '负面' in str(group['sentiment']):
                font_color = '#FF0000'
            elif '正面' in str(group['sentiment']):
                font_color = '#008000'
                
            new_celldata.append({
                'r': start_row,
                'c': 1,
                'v': {
                    'v': group['sentiment'],
                    'm': group['sentiment'],
                    'ct': {'fa': 'General', 't': 'g'},
                    'vt': 1, 'ht': 1,
                    'fc': font_color
                }
            })
            
            # 用户数量
            new_celldata.append({
                'r': start_row,
                'c': 2,
                'v': {
                    'v': group['user_count'],
                    'm': str(group['user_count']),
                    'ct': {'fa': 'General', 't': 'n'},
                    'vt': 1, 'ht': 1
                }
            })
            
            # 用户占比
            new_celldata.append({
                'r': start_row,
                'c': 3,
                'v': {
                    'v': group['user_pct'],
                    'm': group['user_pct'],
                    'ct': {'fa': 'General', 't': 'g'},
                    'vt': 1, 'ht': 1
                }
            })
            
            # 合并单元格配置
            if rows_count > 1:
                for col_idx in range(4):
                    merge_config[f"{start_row}_{col_idx}"] = {
                        "r": start_row,
                        "c": col_idx,
                        "rs": rows_count,
                        "cs": 1
                    }
        
        result = {
            'name': '分析结果',
            'status': 1,  # 设置为活动sheet
            'celldata': new_celldata,
            'config': {
                'merge': merge_config,
                'columnlen': {
                    '0': 200,  # 问题概括
                    '1': 100,  # 用户情绪
                    '2': 70,   # 用户数量
                    '3': 70,   # 用户占比
                    '4': 500   # VOC原声片段
                }
            }
        }
        
        print(f"[Recalculate] Generated {len(new_celldata)} cells with {len(result_list)} groups")
        return jsonify(result), 200
        
    except Exception as e:
        print(f"[Recalculate] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400
    
    # 保存文件
    file_id = str(uuid.uuid4())
    file_path = os.path.join(UPLOAD_FOLDER, f'{file_id}.xlsx')
    file.save(file_path)
    
    # 读取Excel文件并转换为Luckysheet格式
    try:
        wb = load_workbook(file_path)
        sheets_data = []
        
        print(f"[上传] 开始处理文件，共 {len(wb.sheetnames)} 个sheet")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cells = []
            
            print(f"[上传] 处理sheet: {sheet_name}, 最大行: {ws.max_row}, 最大列: {ws.max_column}")
            
            # 读取所有有数据的单元格
            cell_count = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is not None:
                        cells.append({
                            'r': row_idx - 1,
                            'c': col_idx - 1,
                            'v': {
                                'v': cell.value,
                                'm': str(cell.value),
                                'ct': {'fa': 'General', 't': 'g'}
                            }
                        })
                        cell_count += 1
            
            print(f"[上传] Sheet {sheet_name} 共读取 {cell_count} 个单元格")
            
            # 设置列宽
            # 设置列宽
            column = {}
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                width = ws.column_dimensions[col_letter].width if ws.column_dimensions[col_letter].width else 73
                # Luckysheet/FortuneSheet expects key to be string index "0", "1", etc.
                column[str(col_idx - 1)] = int(width) if width else 73
            
            sheet_data = {
                'name': sheet_name,
                'index': len(sheets_data),
                'order': len(sheets_data),
                'status': 1,
                'celldata': cells,
                'config': {
                    'columnlen': column,
                    'rowlen': {}
                },
                'scrollLeft': 0,
                'scrollTop': 0,
                'luckysheet_select_save': [],
                'calc chain': [],
                'isPivotTable': False,
                'pivotTable': {},
                'filter_select': None,
                'filter': None,
                'luckysheet_conditionformat_save': [],
                'frozen': {},
                'chart': [],
                'zoomRatio': 1,
                'image': [],
                'showGridLines': 1,
                'dataVerification': {}
            }
            sheets_data.append(sheet_data)
            print(f"[上传] Sheet {sheet_name} 数据已添加到sheets_data，celldata数量: {len(cells)}")
        
        print(f"[上传] 准备返回数据，共 {len(sheets_data)} 个sheet")
        result = {
            'fileId': file_id,
            'sheets': sheets_data,
            'originalSheets': wb.sheetnames
        }
        print(f"[上传] 返回数据大小: {len(str(result))} 字符")
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': f'处理文件失败: {str(e)}'}), 500

@app.route('/api/analyze', methods=['POST'])
def analyze_voc():
    data = request.json
    file_id = data.get('fileId')
    
    if not file_id:
        return jsonify({'error': '缺少fileId'}), 400
    
    file_path = os.path.join(UPLOAD_FOLDER, f'{file_id}.xlsx')
    if not os.path.exists(file_path):
        return jsonify({'error': '文件不存在'}), 404
    
    # 如果已有任务，先停止它
    if file_id in analysis_tasks:
        analysis_tasks[file_id]['stop_flag'].set()
        # 等待旧任务结束
        analysis_tasks[file_id]['thread'].join(timeout=2)
    
    # 创建停止标志
    stop_flag = threading.Event()
    
    # 创建新任务
    result_container = {'result': None, 'error': None, 'completed': False}
    
    def send_progress(current, total, message):
        """发送进度更新"""
        progress = int((current / total * 100)) if total > 0 else 0
        return f"data: {json.dumps({'type': 'progress', 'current': current, 'total': total, 'progress': progress, 'message': message}, ensure_ascii=False)}\n\n"
    
    def analyze_task(progress_queue):
        try:
            print(f"[分析任务] 开始分析文件: {file_path}")
            # 发送初始进度
            progress_queue.put(('progress', 0, 100, '开始分析...'))
            
            # 设置停止标志到analyzer
            analyzer.set_stop_flag(stop_flag)
            
            # 定义进度回调函数
            def progress_callback(current, total, message):
                if not stop_flag.is_set():
                    print(f"[进度更新] {message} ({current}/{total})")
                    progress_queue.put(('progress', current, total, message))
            
            # 分析VOC数据
            print(f"[分析任务] 调用 analyze_file...")
            analyzer.progress_callback = progress_callback
            analyzed_sheets = analyzer.analyze_file(file_path)
            print(f"[分析任务] 分析完成，得到 {len(analyzed_sheets) if analyzed_sheets else 0} 个sheet")
            
            if stop_flag.is_set():
                print(f"[分析任务] 检测到停止标志")
                result_container['error'] = '分析被用户终止'
                progress_queue.put(('error', '分析被用户终止'))
                return
            
            if not analyzed_sheets:
                print(f"[分析任务] 分析结果为空")
                result_container['error'] = '分析结果为空，请检查文件格式'
                progress_queue.put(('error', '分析结果为空，请检查文件格式'))
                return
            
            result_container['result'] = {
                'fileId': file_id,
                'sheets': analyzed_sheets
            }
            print(f"[分析任务] 发送完成消息，包含 {len(analyzed_sheets)} 个sheet")
            progress_queue.put(('complete', result_container['result']))
        except KeyboardInterrupt:
            print(f"[分析任务] 捕获到 KeyboardInterrupt")
            result_container['error'] = '分析被中断'
            progress_queue.put(('error', '分析被中断'))
        except Exception as e:
            import traceback
            error_detail = str(e)
            print(f"[分析任务] 分析错误详情: {traceback.format_exc()}")
            # 检查是否是用户终止
            if '分析被用户终止' in error_detail or stop_flag.is_set():
                result_container['error'] = '分析被用户终止'
                progress_queue.put(('error', '分析被用户终止'))
            else:
                result_container['error'] = f'分析失败: {error_detail}'
                progress_queue.put(('error', result_container['error']))
        finally:
            result_container['completed'] = True
            progress_queue.put(('done', None))
            print(f"[分析任务] 任务完成，清理资源")
            # 清理任务
            if file_id in analysis_tasks:
                del analysis_tasks[file_id]
    
    # 使用队列来传递进度更新
    progress_queue = queue.Queue()
    
    # 启动分析线程
    thread = threading.Thread(target=analyze_task, args=(progress_queue,))
    thread.daemon = True
    thread.start()
    
    analysis_tasks[file_id] = {
        'stop_flag': stop_flag,
        'thread': thread
    }
    
    # 使用SSE流式响应
    @stream_with_context
    def generate():
        print(f"[SSE] 开始生成流式响应 for file_id: {file_id}")
        timeout = 300  # 5分钟超时
        start_time = time.time()
        
        while True:
            try:
                # 检查超时
                if time.time() - start_time > timeout:
                    print(f"[SSE] 超时")
                    yield f"data: {json.dumps({'type': 'error', 'message': '分析超时'}, ensure_ascii=False)}\n\n"
                    break
                
                # 检查线程是否还在运行
                if not thread.is_alive() and result_container['completed']:
                    # 线程已完成，检查是否有最终结果
                    print(f"[SSE] 线程已完成，检查结果...")
                    if result_container['error']:
                        print(f"[SSE] 发送错误: {result_container['error']}")
                        yield f"data: {json.dumps({'type': 'error', 'message': result_container['error']}, ensure_ascii=False)}\n\n"
                    elif result_container['result']:
                        # 如果线程完成但没有通过队列发送complete消息，直接发送结果
                        print(f"[SSE] 发送结果（线程已完成但未通过队列）")
                        yield f"data: {json.dumps({'type': 'complete', 'data': result_container['result']}, ensure_ascii=False)}\n\n"
                    break
                
                # 从队列获取进度更新（非阻塞）
                try:
                    update_type, *args = progress_queue.get(timeout=0.5)
                    print(f"[SSE] 收到队列消息: {update_type}")
                    
                    if update_type == 'progress':
                        current, total, message = args
                        progress = int((current / total * 100)) if total > 0 else 0
                        print(f"[SSE] 发送进度: {message} ({current}/{total}, {progress}%)")
                        yield f"data: {json.dumps({'type': 'progress', 'current': current, 'total': total, 'progress': progress, 'message': message}, ensure_ascii=False)}\n\n"
                    elif update_type == 'complete':
                        result = args[0]
                        print(f"[SSE] 发送完成消息，包含 {len(result.get('sheets', []))} 个sheet")
                        yield f"data: {json.dumps({'type': 'complete', 'data': result}, ensure_ascii=False)}\n\n"
                        break
                    elif update_type == 'error':
                        error_msg = args[0]
                        print(f"[SSE] 发送错误: {error_msg}")
                        yield f"data: {json.dumps({'type': 'error', 'message': error_msg}, ensure_ascii=False)}\n\n"
                        break
                    elif update_type == 'done':
                        print(f"[SSE] 收到done消息")
                        # 检查是否有结果需要发送
                        if result_container['result']:
                            print(f"[SSE] 发送结果（done消息后）")
                            yield f"data: {json.dumps({'type': 'complete', 'data': result_container['result']}, ensure_ascii=False)}\n\n"
                        break
                except queue.Empty:
                    # 队列为空，继续等待
                    continue
                    
            except Exception as e:
                import traceback
                print(f"[SSE] 生成错误: {e}")
                print(traceback.format_exc())
                yield f"data: {json.dumps({'type': 'error', 'message': f'服务器错误: {str(e)}'}, ensure_ascii=False)}\n\n"
                break
    
    return Response(generate(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'
    })

@app.route('/api/analyze/stop', methods=['POST'])
def stop_analyze():
    data = request.json
    file_id = data.get('fileId')
    
    if not file_id:
        return jsonify({'error': '缺少fileId'}), 400
    
    if file_id in analysis_tasks:
        analysis_tasks[file_id]['stop_flag'].set()
        print(f"[停止分析] 已设置停止标志 for file_id: {file_id}")
        return jsonify({'message': '分析已终止'})
    else:
        return jsonify({'message': '没有正在进行的分析任务'})

if __name__ == '__main__':
    app.run(debug=True, port=5000)

