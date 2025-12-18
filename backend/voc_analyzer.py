import openpyxl
import requests
import json
import re
import os
import math
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 尝试导入配置文件
try:
    from config import HF_API_TOKEN, TONGYI_API_KEY, TONGYI_MODEL, API_PRIORITY
except ImportError:
    # 如果配置文件不存在，使用默认值
    HF_API_TOKEN = None
    TONGYI_API_KEY = None
    TONGYI_MODEL = "qwen-turbo"
    API_PRIORITY = ["hf_token", "tongyi", "hf_free", "local"]

class VOCAnalyzer:
    def __init__(self):
        # 加载API配置
        self.hf_token = HF_API_TOKEN or os.getenv('HF_API_TOKEN')
        self.tongyi_key = TONGYI_API_KEY or os.getenv('TONGYI_API_KEY')
        # 获取模型名称，如果未配置则使用默认值
        self.tongyi_model = TONGYI_MODEL if 'TONGYI_MODEL' in globals() else (os.getenv('TONGYI_MODEL') or "qwen-turbo")
        self.api_priority = API_PRIORITY
        
        # Hugging Face API端点（使用Token）
        self.hf_api_urls = [
            "https://api-inference.huggingface.co/models/Qwen/Qwen2.5-7B-Instruct",
            "https://api-inference.huggingface.co/models/Qwen/Qwen2-7B-Instruct",
            "https://api-inference.huggingface.co/models/Qwen/Qwen2-1.5B-Instruct",
            "https://api-inference.huggingface.co/models/Qwen/Qwen2.5-14B-Instruct",
        ]
        
        # Hugging Face免费API端点（无需Token，但可能不可用）
        self.hf_free_api_urls = [
            "https://api-inference.huggingface.co/models/Qwen/Qwen2.5-7B-Instruct",
            "https://api-inference.huggingface.co/models/Qwen/Qwen2-7B-Instruct",
            "https://api-inference.huggingface.co/models/Qwen/Qwen2-1.5B-Instruct",
        ]
        
        # 通义千问API端点
        self.tongyi_api_url = "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation"
        
        self.current_api_index = 0
        self.use_local_analysis = False
        self.stop_flag = None
        
        # 打印配置信息
        print(f"[VOC Analyzer] 初始化完成")
        if self.hf_token:
            print(f"[VOC Analyzer] Hugging Face Token已配置")
        if self.tongyi_key:
            print(f"[VOC Analyzer] 通义千问API Key已配置，模型: {self.tongyi_model}")
        print(f"[VOC Analyzer] API优先级: {', '.join(self.api_priority)}")
    
    def set_stop_flag(self, stop_flag):
        """设置停止标志"""
        self.stop_flag = stop_flag
    
    def analyze_with_ai(self, text):
        """使用Qwen AI分析文本情感和分类，按优先级尝试不同的API"""
        if self.use_local_analysis:
            return self.local_analyze(text)
        
        # 构造prompt
        prompt = f"""Role (角色设定):
你是一名拥有10年经验的 B2B SaaS 产品体验分析师。你的任务是清洗用户反馈数据（VOC），精准识别用户痛点，并进行标准化的分类归纳。

Critical Rules (核心判别规则 - 必须严格遵守):
1. Bug vs. 灵活性 (最高优先级):
   - 判定为 [功能 - Bug/稳定性]：当用户描述"操作无效"、"报错"、"显示异常"、"死机"、"明明设置了但没反应"等预期功能失效的情况。
   - 判定为 [功能 - 灵活性/配置能力]：只有当用户明确表示"希望能自定义..."、"想要支持...功能"、"目前选项太少"等新增需求时。
   - 案例："主页板块加链接后图片不显示" -> [功能 - Bug/稳定性]。

2. 概括度控制 (归纳法):
   - 将相似的具体问题向上归纳到父类目。
   - 案例："新手教程缺失"、"开发文档不全" -> [服务 - 帮助与引导]。

Taxonomy (标准化分类体系 - 请仅从以下列表中选择):
- 功能 - Bug/稳定性
- 功能 - 灵活性/配置能力
- 功能 - 实用性/完整度
- 体验 - 操作复杂度
- 体验 - 性能/加载速度
- 资源 - 模板丰富度
- 资源 - 插件生态
- 服务 - 帮助与引导

请分析以下用户反馈，返回一个JSON对象：
{{
    "category": "必须从上方Taxonomy列表中选择一个标准的分类名称 (例如: 功能 - Bug/稳定性)",
    "sentiment": "正面😊/负面😠/中性😐",
    "rationale": "简短的分类理由"
}}

用户反馈：{text}

请只返回单个JSON对象："""
        
        # 按优先级尝试不同的API
        for api_type in self.api_priority:
            if api_type == "hf_token" and self.hf_token:
                result = self._try_huggingface_token(prompt, text)
                if result:
                    return result
            elif api_type == "tongyi" and self.tongyi_key:
                result = self._try_tongyi_api(prompt, text)
                if result:
                    return result
            elif api_type == "hf_free":
                result = self._try_huggingface_free(prompt, text)
                if result:
                    return result
            elif api_type == "local":
                print("[Qwen API] 使用本地分析")
                return self.local_analyze(text)
        
        # 所有API都失败，使用本地分析
        print("[Qwen API] 所有API都不可用，使用本地分析")
        return self.local_analyze(text)
    
    def _try_huggingface_token(self, prompt, text):
        """尝试使用Hugging Face API Token"""
        for api_url in self.hf_api_urls:
            try:
                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {self.hf_token}"
                }
                payload = {
                    "inputs": prompt,
                    "parameters": {
                        "max_new_tokens": 150,
                        "temperature": 0.3,
                        "return_full_text": False
                    }
                }
                
                print(f"[HF Token API] 尝试调用: {api_url}")
                response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    result = response.json()
                    print(f"[HF Token API] 调用成功")
                    return self.parse_ai_result(result, text)
                elif response.status_code == 503:
                    error_info = response.json() if response.content else {}
                    estimated_time = error_info.get('estimated_time', 0)
                    print(f"[HF Token API] 模型正在加载，预计等待时间: {estimated_time}秒")
                    if estimated_time and estimated_time < 30:
                        import time
                        time.sleep(min(estimated_time + 2, 30))
                        retry_response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                        if retry_response.status_code == 200:
                            return self.parse_ai_result(retry_response.json(), text)
                    continue
                else:
                    print(f"[HF Token API] 错误 {response.status_code}: {response.text[:200]}")
                    continue
            except Exception as e:
                print(f"[HF Token API] 调用失败: {e}")
                continue
        return None
    
    def _try_tongyi_api(self, prompt, text):
        """尝试使用通义千问API"""
        try:
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.tongyi_key}"
            }
            payload = {
                "model": self.tongyi_model,  # 使用配置的模型
                "input": {
                    "messages": [
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ]
                },
                "parameters": {
                    "max_tokens": 150,
                    "temperature": 0.3
                }
            }
            
            print(f"[通义千问API] 尝试调用模型: {self.tongyi_model}")
            response = requests.post(self.tongyi_api_url, headers=headers, json=payload, timeout=30)
            
            if response.status_code == 200:
                result = response.json()
                print(f"[通义千问API] 响应状态: 200")
                
                # 通义千问API的响应格式可能是两种：
                # 1. 新格式: result['output']['text'] 直接包含文本
                # 2. 旧格式: result['output']['choices'][0]['message']['content']
                generated_text = None
                
                if result.get('output'):
                    output = result['output']
                    # 尝试新格式（text字段）
                    if 'text' in output:
                        generated_text = output['text']
                        print(f"[通义千问API] 使用text字段获取结果")
                    # 尝试旧格式（choices字段）
                    elif 'choices' in output and len(output['choices']) > 0:
                        generated_text = output['choices'][0]['message']['content']
                        print(f"[通义千问API] 使用choices字段获取结果")
                
                if generated_text:
                    print(f"[通义千问API] 调用成功，返回文本长度: {len(generated_text)}")
                    # 解析结果
                    return self.parse_ai_result({'generated_text': generated_text}, text)
                else:
                    print(f"[通义千问API] 响应格式异常，未找到text或choices: {result}")
                    return None
            elif response.status_code == 429:
                # 速率限制，等待后重试
                error_info = response.json() if response.content else {}
                wait_time = 2  # 默认等待2秒
                print(f"[通义千问API] 速率限制(429)，等待{wait_time}秒后重试...")
                import time
                time.sleep(wait_time)
                # 重试一次
                retry_response = requests.post(self.tongyi_api_url, headers=headers, json=payload, timeout=30)
                if retry_response.status_code == 200:
                    result = retry_response.json()
                    if result.get('output'):
                        output = result['output']
                        if 'text' in output:
                            generated_text = output['text']
                        elif 'choices' in output and len(output['choices']) > 0:
                            generated_text = output['choices'][0]['message']['content']
                        else:
                            generated_text = None
                        
                        if generated_text:
                            print(f"[通义千问API] 重试成功")
                            return self.parse_ai_result({'generated_text': generated_text}, text)
                print(f"[通义千问API] 重试后仍失败，返回None以尝试下一个API")
                return None
            else:
                print(f"[通义千问API] 错误 {response.status_code}: {response.text[:200]}")
                return None
        except Exception as e:
            print(f"[通义千问API] 调用失败: {e}")
            import traceback
            print(f"[通义千问API] 错误详情: {traceback.format_exc()}")
            return None
    
    def _try_huggingface_free(self, prompt, text):
        """尝试使用Hugging Face免费API（无需Token）"""
        for api_url in self.hf_free_api_urls:
            try:
                headers = {"Content-Type": "application/json"}
                payload = {
                    "inputs": prompt,
                    "parameters": {
                        "max_new_tokens": 150,
                        "temperature": 0.3,
                        "return_full_text": False
                    }
                }
                
                print(f"[HF Free API] 尝试调用: {api_url}")
                response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                
                if response.status_code == 200:
                    result = response.json()
                    print(f"[HF Free API] 调用成功")
                    return self.parse_ai_result(result, text)
                elif response.status_code == 503:
                    error_info = response.json() if response.content else {}
                    estimated_time = error_info.get('estimated_time', 0)
                    print(f"[HF Free API] 模型正在加载，预计等待时间: {estimated_time}秒")
                    if estimated_time and estimated_time < 30:
                        import time
                        time.sleep(min(estimated_time + 2, 30))
                        retry_response = requests.post(api_url, headers=headers, json=payload, timeout=30)
                        if retry_response.status_code == 200:
                            return self.parse_ai_result(retry_response.json(), text)
                    continue
                elif response.status_code == 410:
                    print(f"[HF Free API] 模型不可用(410 - Gone)")
                    continue
                elif response.status_code == 429:
                    print(f"[HF Free API] 请求过多(429)")
                    import time
                    time.sleep(2)
                    continue
                else:
                    print(f"[HF Free API] 错误 {response.status_code}: {response.text[:200]}")
                    continue
            except Exception as e:
                print(f"[HF Free API] 调用失败: {e}")
                continue
        return None
    
    def local_analyze(self, text):
        """本地规则分析（备用方案）"""
        text_lower = text.lower()
        
        # 情感分析关键词（更全面的中文关键词）
        positive_keywords = ['好', '满意', '喜欢', '推荐', '优秀', '棒', '赞', '不错', '很好', '完美', 
                            '赞', '给力', '好用', '方便', '快捷', '流畅', '清晰', '美观', '实用', 
                            '贴心', '专业', '高效', '稳定', '可靠', '值得', '超值', '惊喜']
        negative_keywords = ['差', '不好', '失望', '问题', '错误', '慢', '卡', '崩溃', 'bug', '故障',
                            '糟糕', '垃圾', '难用', '复杂', '麻烦', '延迟', '卡顿', '闪退', '死机',
                            '不兼容', '缺失', '不足', '缺陷', '漏洞', '不安全', '贵', '不值']
        
        positive_count = sum(1 for kw in positive_keywords if kw in text)
        negative_count = sum(1 for kw in negative_keywords if kw in text)
        
        # 判断情感
        if positive_count > negative_count and positive_count > 0:
            sentiment = '正面'
        elif negative_count > 0:
            sentiment = '负面'
        else:
            sentiment = '中性'
        
        # 简单分类
        summary = self.categorize_text(text)
        
        # 添加简单表情
        sentiment_emoji = {
            '正面': '正面😊',
            '负面': '负面😠',
            '中性': '中性😐'
        }
        
        return [{
            'sentiment': sentiment_emoji.get(sentiment, sentiment),
            'summary': summary,
            'snippet': text,
            'confidence': 0.7
        }]
    
    def categorize_text(self, text):
        """简单的文本分类"""
        text_lower = text.lower()
        
        categories = {
            '功能 - Bug/稳定性': ['功能', '不能', '无法', '不支持', '缺少', '没有', '缺失', '不完善', '不完整', '死机', '报错', '失效', '不显示'],
            '功能 - 灵活性/配置能力': ['自定义', '配置', '选项', '灵活', '更多功能', '支持', '设置'],
            '功能 - 实用性/完整度': ['半成品', '不好用', '鸡肋', '没用', '奇怪'],
            '体验 - 操作复杂度': ['难找', '步骤', '复杂', '麻烦', '逻辑', '反人类', '难用'],
            '体验 - 性能/加载速度': ['慢', '卡', '延迟', '加载', '响应', '卡顿', '速度', '性能', '优化'],
            '资源 - 模板丰富度': ['模板', '风格', '主题', '样式'],
            '资源 - 插件生态': ['插件', '扩展', '应用'],
            '服务 - 帮助与引导': ['文档', '教程', '指引', '说明', '帮助', '客服', '支持'],
        }
        
        # 计算每个类别的匹配分数
        category_scores = {}
        for category, keywords in categories.items():
            score = sum(1 for kw in keywords if kw in text)
            if score > 0:
                category_scores[category] = score
        
        # 返回得分最高的类别
        if category_scores:
            return max(category_scores.items(), key=lambda x: x[1])[0]
        
        return '其他问题'
    
    def parse_ai_result(self, result, text):
        """解析AI返回的JSON结果"""
        import json
        try:
            generated_text = ""
            # 获取生成的文本
            if isinstance(result, dict):
                if 'generated_text' in result:
                    generated_text = result['generated_text']
                elif 'text' in result:
                     generated_text = result['text']
            elif isinstance(result, str):
                generated_text = result

            if not generated_text:
                return None

            # 尝试解析JSON
            # 清理可能的markdown标记
            clean_text = generated_text.strip()
            if clean_text.startswith('```json'):
                clean_text = clean_text[7:]
            if clean_text.startswith('```'):
                clean_text = clean_text[3:]
            if clean_text.endswith('```'):
                clean_text = clean_text[:-3]
            clean_text = clean_text.strip()
            
            # 找到JSON数组部分
            start = clean_text.find('[')
            end = clean_text.rfind(']') + 1
            
            parsed = None
            if start != -1 and end != -1:
                json_str = clean_text[start:end]
                try:
                    parsed = json.loads(json_str)
                except:
                    pass
            
            # 如果没找到数组，尝试解析整个文本为对象
            if not parsed:
                 try:
                    parsed = json.loads(clean_text)
                    if isinstance(parsed, dict):
                        parsed = [parsed]
                 except:
                    pass

            if not parsed:
                return None
                
            validated_results = []
            for item in parsed:
                sentiment = item.get('sentiment', '中性😐')
                # 适配新Prompt的返回字段 'category'
                summary = item.get('category') or item.get('summary') or '其他问题'
                snippet = item.get('snippet', text) 
                
                validated_results.append({
                    'sentiment': sentiment,
                    'summary': summary,
                    'snippet': snippet,
                    'confidence': 0.85
                })
            
            return validated_results

        except Exception as e:
            print(f"[Parse] Error: {str(e)}")
            return None
            
    def analyze_and_categorize(self, rows_data, feedback_col):
        """分析并分类数据（支持多观点拆分）"""
        print(f"[Analyze] Analyzing {len(rows_data)} rows...")
        
        # 扁平化的所有意见列表，包含 row_id 用于计算用户数
        all_opinions = []
        
        total_rows = len(rows_data)
        if hasattr(self, 'progress_callback') and self.progress_callback:
            self.progress_callback(0, total_rows, f'开始分析，共 {total_rows} 条反馈...')
            
        for idx, row_info in enumerate(rows_data, 1):
            if self.stop_flag and self.stop_flag.is_set():
                raise KeyboardInterrupt("分析被用户终止")
                
            if hasattr(self, 'progress_callback') and self.progress_callback:
                self.progress_callback(idx, total_rows, f'正在分析第 {idx}/{total_rows} 条反馈...')
            
            # AI 分析返回列表
            analysis_list = self.analyze_with_ai(row_info[feedback_col])
            
            # API 延迟
            if self.tongyi_key and idx < total_rows:
                import time
                time.sleep(0.3)
            
            # 扁平化存储 (不拆分，直接存)
            # 兼容返回列表的情况（如果有）
            first_opinion = analysis_list[0] if analysis_list and len(analysis_list) > 0 else {
                'summary': '其他问题', 'sentiment': '中性😐'
            }

            
            all_opinions.append({
                'row_id': idx,
                'summary': first_opinion['summary'],
                'sentiment': first_opinion['sentiment'],
                'snippet': row_info[feedback_col], # snippet直接等于全文
                'full_feedback': row_info[feedback_col],
                'row_data': row_info
            })
                
        return all_opinions

    def generate_analysis_sheet(self, all_opinions, total_users, sheet_name, sort_by='user', original_columns=None):
        """生成归类后的分析Sheet (包含原始列)
        - 将同类VOC行放在一起，并为分组创建合并的总问题标题
        - 功能/体验等分类单独放一列，不与问题标题混在一起
        """
        if original_columns is None:
            original_columns = []

        # 先按“问题标题”分组，保持出现顺序
        grouped = []
        group_map = {}

        def split_summary(summary_text):
            """拆分分类：前半部分为归类（功能/体验），后半部分为总问题标题"""
            if not summary_text:
                return "其他问题", ""
            text = str(summary_text).strip()
            parts = [p.strip() for p in re.split(r'[-—]', text, maxsplit=1)]
            if len(parts) == 2 and parts[0] and parts[1]:
                return parts[1], parts[0]  # (问题标题, 问题归类)
            return text, ""  # 没有明确归类时，整句作为标题

        for opinion in all_opinions:
            title, category = split_summary(opinion.get('summary'))
            key = (title, category)
            if key not in group_map:
                group_map[key] = []
                grouped.append((key, group_map[key]))
            group_map[key].append({**opinion, 'title': title, 'category': category})

        # 构建Sheet Data
        celldata = []

        # 表头: [问题总标题, 问题归类, 用户情绪] + 原始列
        headers = ['问题总标题', '问题归类', '用户情绪'] + original_columns

        for col_idx, header in enumerate(headers):
            celldata.append({
                'r': 0,
                'c': col_idx,
                'v': {
                    'v': header,
                    'm': header,
                    'ct': {'fa': 'General', 't': 'g'},
                    'bg': '#EDEBE9',
                    'bl': 1
                }
            })

        current_row = 1
        config = {'merge': {}, 'columnlen': {}}

        # 按分组填充数据，并对分组列做合并
        for (title, category), opinions in grouped:
            start_row = current_row
            group_rows = len(opinions)

            for opinion in opinions:
                row_idx = current_row

                # 问题总标题 & 问题归类 & 用户情绪（只在组首生成，之后依赖合并）
                if row_idx == start_row:
                    celldata.append({
                        'r': row_idx,
                        'c': 0,
                        'v': {
                            'v': title,
                            'm': title,
                            'ct': {'fa': 'General', 't': 'g'},
                            'vt': 1,
                            'ht': 1,
                            'bg': '#F6F8FA'
                        }
                    })

                    celldata.append({
                        'r': row_idx,
                        'c': 1,
                        'v': {
                            'v': category or '未分类',
                            'm': category or '未分类',
                            'ct': {'fa': 'General', 't': 'g'},
                            'vt': 1,
                            'ht': 1,
                            'bg': '#F6F8FA'
                        }
                    })

                    font_color = '#000000'
                    if '负面' in str(opinion['sentiment']):
                        font_color = '#FF0000'
                    elif '正面' in str(opinion['sentiment']):
                        font_color = '#008000'

                    celldata.append({
                        'r': row_idx,
                        'c': 2,
                        'v': {
                            'v': opinion['sentiment'],
                            'm': opinion['sentiment'],
                            'ct': {'fa': 'General', 't': 'g'},
                            'fc': font_color,
                            'vt': 1,
                            'ht': 1
                        }
                    })

                # 原始列数据（从列3开始）
                for col_i, col_name in enumerate(original_columns):
                    val = opinion['row_data'].get(col_name, '')
                    if isinstance(val, float):
                        import math
                        if math.isnan(val) or math.isinf(val):
                            val = ""
                    val_str = str(val)

                    celldata.append({
                        'r': row_idx,
                        'c': 3 + col_i,
                        'v': {
                            'v': val_str,
                            'm': val_str,
                            'ct': {'fa': 'General', 't': 'g'}
                        }
                    })

                current_row += 1

            # 生成合并配置（将同组的“问题总标题”“问题归类”“用户情绪”列合并）
            if group_rows > 1:
                for col_idx in (0, 1, 2):
                    config['merge'][f"{start_row}_{col_idx}"] = {
                        "r": start_row,
                        "c": col_idx,
                        "rs": group_rows,
                        "cs": 1
                    }

        # 列宽配置
        column_len = {
            '0': 220,  # 问题总标题
            '1': 120,  # 问题归类（功能/体验等）
            '2': 100,  # 用户情绪
        }

        return {
            'name': sheet_name,
            'celldata': celldata,
            'config': {
                'merge': config.get('merge', {}),
                'columnlen': column_len
            }
        }

    def create_sheet_data(self, ws, sheet_name, sheet_idx):
        """将Worksheet转换为Luckysheet格式的数据"""
        celldata = []
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 读取所有单元格
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    # 处理各种类型的值，确保不会产生NaN
                    try:
                        raw_value = cell.value
                        cell_value = str(raw_value)
                        
                        # 检查特殊值
                        lower_val = cell_value.lower()
                        if lower_val == 'nan' or lower_val == 'inf' or lower_val == '-inf':
                            cell_value = ""
                    except Exception:
                        cell_value = ""
                    
                    celldata.append({
                        "r": row - 1,
                        "c": col - 1,
                        "v": {
                            "v": cell_value,
                            "m": cell_value,
                            "ct": {"fa": "General", "t": "g"}
                        }
                    })
        
        return {
            "name": sheet_name,
            "index": str(sheet_idx),
            "order": sheet_idx,
            "status": 1 if sheet_idx == 0 else 0,
            "celldata": celldata
        }

    def analyze_dataframe(self, df, original_sheet_data=None):
        """分析DataFrame的核心逻辑
        
        Args:
            df: pandas DataFrame containing the data to analyze
            original_sheet_data: Optional dict for original sheet (if None, will be generated from df)
        
        Returns:
            list of sheet data dicts
        """
        try:
            columns = df.columns.tolist()
            
            # 智能识别反馈列
            feedback_col = None
            
            # 1. 关键词匹配 (优先级最高)
            keywords = ['feedback', 'comment', 'content', 'voice', 'opinion', '建议', '反馈', '意见', '原声', '内容', '评价']
            for col in columns:
                if any(k in str(col).lower() for k in keywords):
                    feedback_col = col
                    print(f"[Analyze] Automatically detected feedback column by keyword: {feedback_col}")
                    break
            
            # 2. 如果没找到，使用内容平均长度判断
            if not feedback_col:
                max_avg_len = 0
                best_col = columns[0]
                
                for col in columns:
                    sample_texts = [str(x) for x in df[col].head(10).tolist() if pd.notna(x)]
                    if not sample_texts:
                        continue
                        
                    avg_len = sum(len(t) for t in sample_texts) / len(sample_texts)
                    
                    if avg_len > max_avg_len:
                        max_avg_len = avg_len
                        best_col = col
                
                feedback_col = best_col
                print(f"[Analyze] Automatically detected feedback column by length: {feedback_col} (Avg Len: {max_avg_len:.1f})")

            print(f"[Analyze] Using column '{feedback_col}' as feedback source.")
            rows = df.to_dict('records')
            total_users = len(rows)
            
            # 分析并获取扁平化数据
            all_opinions = self.analyze_and_categorize(rows, feedback_col)
            
            sheets_data = []
            
            # 添加原始数据Sheet
            if original_sheet_data:
                sheets_data.append(original_sheet_data)
            else:
                # 从DataFrame生成原始数据sheet
                original_sheet = self._dataframe_to_sheet_data(df, "原始数据", 0)
                sheets_data.append(original_sheet)

            # 生成分析结果 Sheet
            sheet_user = self.generate_analysis_sheet(all_opinions, total_users, "分析结果", 'user', original_columns=columns)
            sheet_user['index'] = 1
            sheet_user['order'] = 1
            sheet_user['status'] = 1
            sheets_data.append(sheet_user)
            
            return sheets_data
            
        except Exception as e:
            print(f"[Analyze] Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
    
    def _dataframe_to_sheet_data(self, df, sheet_name, sheet_idx):
        """将DataFrame转换为sheet data格式"""
        celldata = []
        
        # 表头
        for col_idx, col_name in enumerate(df.columns):
            celldata.append({
                'r': 0,
                'c': col_idx,
                'v': {
                    'v': str(col_name),
                    'm': str(col_name),
                    'ct': {'fa': 'General', 't': 'g'}
                }
            })
        
        # 数据行
        for row_idx, row in df.iterrows():
            for col_idx, col_name in enumerate(df.columns):
                val = row[col_name]
                if pd.notna(val):
                    val_str = str(val)
                    celldata.append({
                        'r': row_idx + 1,
                        'c': col_idx,
                        'v': {
                            'v': val_str,
                            'm': val_str,
                            'ct': {'fa': 'General', 't': 'g'}
                        }
                    })
        
        return {
            'name': sheet_name,
            'index': str(sheet_idx),
            'order': sheet_idx,
            'status': 1 if sheet_idx == 0 else 0,
            'celldata': celldata
        }

    def analyze_file(self, filepath):
        """分析文件的主入口"""
        try:
            print(f"[Analyze] Reading file: {filepath}")
            if filepath.endswith('.csv'):
                df = pd.read_csv(filepath)
            else:
                df = pd.read_excel(filepath)
            
            # 为了保持兼容性，使用 openpyxl 读取生成原始 sheet data
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            original_sheet = self.create_sheet_data(ws, "原始数据", 0)
            
            # 调用核心分析逻辑
            return self.analyze_dataframe(df, original_sheet_data=original_sheet)
            
        except Exception as e:
            print(f"[Analyze] Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return []
